import re
import openpyxl

def extract_keys():
    try:
        with open("cardsData.js", "r", encoding="utf-8") as f:
            content = f.read()
            
        # Regex to find all property names (keys) inside objects
        # Looking for "key": or 'key': patterns
        keys = re.findall(r'["\']([^"\']+)["\']\s*:', content)
        
        # Unique keys, preserving order of appearance
        seen = set()
        unique_keys = []
        for k in keys:
            if k not in seen:
                seen.add(k)
                unique_keys.append(k)
        
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Property Names (Keys)"
        
        ws["A1"] = "Sr. No."
        ws["B1"] = "Property Name (Key)"
        
        # Style header
        for cell in ws["1:1"]:
            cell.font = openpyxl.styles.Font(bold=True)
            
        for i, key in enumerate(unique_keys, start=1):
            ws.cell(row=i+1, column=1, value=i)
            ws.cell(row=i+1, column=2, value=key)
            
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 40
        
        output_file = "all_property_names.xlsx"
        wb.save(output_file)
        print(f"Successfully extracted {len(unique_keys)} unique property names to {output_file}")
        return output_file
    except Exception as e:
        print(f"Error: {e}")
        return None

if __name__ == "__main__":
    extract_keys()
