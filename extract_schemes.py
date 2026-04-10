import re
import openpyxl

with open("cardsData.js", "r", encoding="utf-8") as f:
    content = f.read()

# Extract all schemeName values
schemes = re.findall(r'"schemeName"\s*:\s*"([^"]+)"', content)

# Unique, preserving order
seen = set()
unique_schemes = []
for s in schemes:
    if s not in seen:
        seen.add(s)
        unique_schemes.append(s)

# Write to Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Scheme Names"

ws["A1"] = "Sr. No."
ws["B1"] = "Scheme Name"

# Style header
for cell in ws["1:1"]:
    cell.font = openpyxl.styles.Font(bold=True)

for i, name in enumerate(unique_schemes, start=1):
    ws.cell(row=i+1, column=1, value=i)
    ws.cell(row=i+1, column=2, value=name)

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 40

wb.save("unique_schemes.xlsx")
print(f"Done! {len(unique_schemes)} unique schemes saved to unique_schemes.xlsx")
